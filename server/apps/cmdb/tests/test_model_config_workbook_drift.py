"""模型工作簿变更边界。

基线提交 a8b68bce46d562358e667441f57aae12330d4fa7 之后，本任务只允许：
1. 注册 Consul 模型并新增 attr-consul；
2. Redis 新增 topo_mode/cluster_uuid/slaves/master；
3. Docker 新增 status。
"""

import hashlib
import json
import zipfile
from pathlib import Path

import openpyxl


MODEL_CONFIG = Path(__file__).parents[1] / "support-files" / "model_config.xlsx"
BASE_SHEET_NAMES_SHA256 = "498d8c4dad4d3f966b31556f88791e4e49bf4dd9d480f32c4504ba7a86a4c477"
BASE_MODELS_SHA256 = "8f1978db9443074f0b2cbfeb25d43c24ba30a6c5398cd8f2dc2ba998d9e5c331"
BASE_REDIS_NONBLANK_SHA256 = "9523028338227deae114d84e8e536d259c140e7145a569eb4c8aeb0f41e431f5"
BASE_DOCKER_NONBLANK_SHA256 = "b105a9a2a0b857ee0596ff1cb908b01dc7184c7de54f412134f52a0e3a51573c"
BASE_UNTOUCHED_SHEETS_SHA256 = "b61f193f41f3d885bdc95cc3f91961a2574c5458384009d95b375524e6f2d64a"

REDIS_ADDITIONS = {
    "topo_mode": (
        "拓扑模式",
        "str",
        '{"validation_type":"unrestricted","custom_regex":"","widget_type":"single_line"}',
        "技术信息",
        False,
        True,
        False,
        None,
        "",
        False,
        "occasional",
    ),
    "cluster_uuid": (
        "集群标识",
        "str",
        '{"validation_type":"unrestricted","custom_regex":"","widget_type":"single_line"}',
        "技术信息",
        False,
        True,
        False,
        None,
        "",
        False,
        "occasional",
    ),
    "slaves": (
        "从节点列表",
        "str",
        '{"validation_type":"unrestricted","custom_regex":"","widget_type":"single_line"}',
        "技术信息",
        False,
        True,
        False,
        None,
        "",
        False,
        "occasional",
    ),
    "master": (
        "主节点地址",
        "str",
        '{"validation_type":"unrestricted","custom_regex":"","widget_type":"single_line"}',
        "技术信息",
        False,
        True,
        False,
        None,
        "",
        False,
        "occasional",
    ),
}
DOCKER_ADDITION = (
    "容器状态",
    "str",
    '{"validation_type":"unrestricted","custom_regex":"","widget_type":"single_line"}',
    "基本信息",
    False,
    True,
    False,
    None,
    "",
    False,
    "occasional",
)
CONSUL_ATTR_IDS = [
    "inst_name",
    "organization",
    "ip_addr",
    "port",
    "tag",
    "version",
    "install_path",
    "data_dir",
    "conf_path",
    "role",
    "operator",
    "bak_operator",
    "auto_collect",
    "collect_time",
    "collect_task",
]
CONSUL_FRESHNESS = {
    "inst_name": "",
    "organization": "",
    "ip_addr": "occasional",
    "port": "occasional",
    "tag": "",
    "version": "occasional",
    "install_path": "occasional",
    "data_dir": "occasional",
    "conf_path": "occasional",
    "role": "occasional",
    "operator": "occasional",
    "bak_operator": "occasional",
    "auto_collect": "",
    "collect_time": "",
    "collect_task": "",
}


def _digest(value) -> str:
    payload = json.dumps(value, ensure_ascii=False, separators=(",", ":")).encode()
    return hashlib.sha256(payload).hexdigest()


def _rows(workbook, sheet_name: str) -> list[list]:
    return [list(row) for row in workbook[sheet_name].iter_rows(values_only=True)]


def _nonblank_rows(rows: list[list]) -> list[list]:
    return [row for row in rows if any(value is not None for value in row)]


def test_未授权工作表_xml_与基线逐字节一致():
    aggregate = hashlib.sha256()
    with zipfile.ZipFile(MODEL_CONFIG) as workbook:
        for number in range(1, 348):
            if number in {2, 63, 99}:
                continue
            name = f"xl/worksheets/sheet{number}.xml"
            aggregate.update(name.encode())
            aggregate.update(b"\0")
            aggregate.update(workbook.read(name))
            aggregate.update(b"\0")

    assert aggregate.hexdigest() == BASE_UNTOUCHED_SHEETS_SHA256


def test_工作表集合只新增_consul():
    workbook = openpyxl.load_workbook(MODEL_CONFIG, read_only=True, data_only=False)

    assert workbook.sheetnames[-1] == "attr-consul"
    assert len(workbook.sheetnames) == 348
    assert _digest(workbook.sheetnames[:-1]) == BASE_SHEET_NAMES_SHA256


def test_models_只追加_consul():
    workbook = openpyxl.load_workbook(MODEL_CONFIG, read_only=True, data_only=False)
    rows = _rows(workbook, "models")

    assert _digest(rows[:-1]) == BASE_MODELS_SHA256
    assert rows[-1][:4] == ["consul", "Consul", "cc-consul_Consul", "middleware"]
    assert all(value is None for value in rows[-1][4:])


def test_redis_只新增批准字段():
    workbook = openpyxl.load_workbook(MODEL_CONFIG, read_only=True, data_only=False)
    rows = _nonblank_rows(_rows(workbook, "attr-redis"))
    additions = {row[0]: tuple(row[1:]) for row in rows if row[0] in REDIS_ADDITIONS}
    baseline_rows = [row for row in rows if row[0] not in REDIS_ADDITIONS]

    assert additions == REDIS_ADDITIONS
    assert _digest(baseline_rows) == BASE_REDIS_NONBLANK_SHA256


def test_docker_只新增批准字段():
    workbook = openpyxl.load_workbook(MODEL_CONFIG, read_only=True, data_only=False)
    rows = _nonblank_rows(_rows(workbook, "attr-docker"))
    additions = [row for row in rows if row[0] == "status"]
    baseline_rows = [row for row in rows if row[0] != "status"]

    assert len(additions) == 1
    assert tuple(additions[0][1:]) == DOCKER_ADDITION
    assert _digest(baseline_rows) == BASE_DOCKER_NONBLANK_SHA256


def test_consul_字段集合完整且无额外行():
    workbook = openpyxl.load_workbook(MODEL_CONFIG, read_only=True, data_only=False)
    rows = _nonblank_rows(_rows(workbook, "attr-consul"))

    assert rows[0][:8] == ["英文名", "名称", "类型", "数据配置", "分组", "是否唯一", "是否可编辑", "是否必填"]
    assert rows[1][:8] == ["attr_id", "attr_name", "attr_type", "option", "attr_group", "is_only", "editable", "is_required"]
    assert [row[0] for row in rows[2:]] == CONSUL_ATTR_IDS
    expected_prompts = {attr_id: "" for attr_id in CONSUL_ATTR_IDS}
    expected_prompts["tag"] = None
    assert {row[0]: row[9] for row in rows[2:]} == expected_prompts
    assert {row[0]: row[11] for row in rows[2:]} == CONSUL_FRESHNESS


def test_新增字段不含错误占位值():
    workbook = openpyxl.load_workbook(MODEL_CONFIG, read_only=True, data_only=False)

    for sheet_name in ("attr-redis", "attr-docker", "attr-consul"):
        assert all(value != "684" for row in _rows(workbook, sheet_name) for value in row)
