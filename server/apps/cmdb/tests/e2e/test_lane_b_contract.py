import json
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
from unittest import mock

import openpyxl
import pytest
from jsonschema import ValidationError, validate

from apps.cmdb.collection.collect_plugin.base import CollectBase
from apps.cmdb.collection.plugins import get_collection_plugin
from apps.cmdb.tests.e2e.contract_loader import audit_lane_a_evidence
from apps.cmdb.tests.e2e.lane_b_loader import (
    LaneBValidationError,
    build_case_cmdb_schema,
    build_case_vm_schema,
    build_vm_response_from_line_protocol,
    lane_b_entries,
    lane_b_incomplete,
    load_lane_b_evidence,
    parse_model_field_definitions,
    parse_model_field_rows,
)


MODEL_CONFIG = Path(__file__).parents[2] / "support-files" / "model_config.xlsx"


def _fake_task(entry, vm_response):
    first_metric = vm_response["data"]["result"][0]["metric"]
    source_inst_name = first_metric.get("inst_name")
    return SimpleNamespace(
        id=33011,
        model_id=entry.supported_model_id,
        instances=[
            {
                "_id": 1,
                "model_id": entry.supported_model_id,
                "inst_name": source_inst_name
                or f"{entry.supported_model_id}-contract",
                "ip_addr": "192.0.2.10",
                "cloud": "contract-cloud",
                "cloud_name": "contract-cloud",
            }
        ],
        params={},
        is_network_topo=False,
        topology_contract={},
    )


def _run_real_plugin(
    entry,
    vm_response,
    monkeypatch,
    applied_ip_rows=None,
    *,
    clock_offset_seconds=3600,
):
    task = _fake_task(entry, vm_response)
    first_metric = vm_response["data"]["result"][0]["metric"]
    vm_instance_id = str(first_metric.get("instance_id", ""))
    task_id = vm_instance_id.split("_", 1)[1] if "_" in vm_instance_id else "33011"
    monkeypatch.setattr(CollectBase, "get_collect_inst", lambda self: task)
    plugin_cls = get_collection_plugin(entry.task_type, entry.supported_model_id)
    if entry.case_id == "network":
        monkeypatch.setattr(plugin_cls, "get_oid_map", lambda self: {})
    if entry.case_id == "ip":
        monkeypatch.setattr(
            "apps.cmdb.services.ipam_discovery.apply_ip_discovery_vm_rows",
            lambda collect_task, rows: applied_ip_rows.update(
                {"task": collect_task, "rows": rows}
            )
            or {
                "format_data": {
                    "add": [],
                    "update": [],
                    "delete": [],
                    "association": [],
                    "all": len(rows),
                }
            },
        )
    plugin = plugin_cls(
        inst_name=task.instances[0]["inst_name"],
        inst_id=task.instances[0]["_id"],
        task_id=task_id,
    )
    evidence_timestamp = max(
        row["value"][0] for row in vm_response["data"]["result"]
    )

    class FrozenDateTime(datetime):
        @classmethod
        def now(cls, tz=None):
            return cls.fromtimestamp(
                evidence_timestamp + clock_offset_seconds, tz=tz
            )

    # 只冻结墙上时钟；所有插件仍执行生产 timestamp_gt_one_day_ago。
    monkeypatch.setattr(
        "apps.cmdb.collection.collect_util.datetime", FrozenDateTime
    )
    with mock.patch(
        "apps.cmdb.collection.query_vm.Collection.query", return_value=vm_response
    ) as query:
        result = plugin.run()
    return plugin, result, query


def _output_model_id(entry, expected):
    return expected.get("model_id", entry.emitted_model_id)


def _target_mapping_fields(plugin, model_id):
    if model_id != "interface" and hasattr(plugin, "device_map"):
        mapping = plugin.device_map
        assert isinstance(mapping, dict)
        return set(mapping)
    mapping = getattr(plugin, "model_field_mapping", {})
    if isinstance(mapping, dict) and model_id in mapping:
        mapping = mapping[model_id]
    if not isinstance(mapping, dict):
        mapping = getattr(plugin, "field_mapping", {})
    assert isinstance(mapping, dict), f"{model_id}: 无法反射生产 mapping"
    return set(mapping)


def _assert_query_identity(plugin, query, vm_response):
    query.assert_called_once_with(plugin.prom_sql())
    selectors = plugin.prom_sql().split(" or ")
    assert selectors == [
        f"{metric}{{instance_id='{plugin._instance_id}'}}"
        for metric in plugin._metrics
    ]
    assert {
        row["metric"]["instance_id"]
        for row in vm_response["data"]["result"]
    } == {plugin._instance_id}


def _production_model_fields(model_id):
    workbook = openpyxl.load_workbook(MODEL_CONFIG, read_only=True, data_only=True)
    try:
        sheet_name = f"attr-{model_id}"
        assert sheet_name in workbook.sheetnames, f"{model_id}: 生产模型表缺少 {sheet_name}"
        return parse_model_field_rows(
            workbook[sheet_name].iter_rows(values_only=True),
            model_id=model_id,
        )
    finally:
        workbook.close()


def _production_model_field_definitions(model_id):
    workbook = openpyxl.load_workbook(MODEL_CONFIG, read_only=True, data_only=True)
    try:
        sheet_name = f"attr-{model_id}"
        assert sheet_name in workbook.sheetnames, f"{model_id}: 生产模型表缺少 {sheet_name}"
        return parse_model_field_definitions(
            workbook[sheet_name].iter_rows(values_only=True),
            model_id=model_id,
        )
    finally:
        workbook.close()


def _value_matches_model_type(value, attr_type):
    if attr_type == "bool":
        return isinstance(value, bool)
    if attr_type == "int":
        return isinstance(value, int) and not isinstance(value, bool)
    if attr_type == "float":
        return isinstance(value, (int, float)) and not isinstance(value, bool)
    if attr_type in {"str", "time"}:
        return isinstance(value, str)
    return True


def test_生产模型反射允许artifact_tool导出的空短尾行():
    rows = (
        ("模型字段",),
        ("attr_id", "attr_type", "is_required", "option"),
        ("inst_name", "str", True, None),
        (),
        (None,),
    )
    assert parse_model_field_rows(rows, model_id="sample") == {
        "inst_name": "str"
    }


def test_生产模型反射拒绝非空短行而不静默吞格式错误():
    rows = (
        ("模型字段",),
        ("attr_id", "unused", "attr_type", "is_required", "option"),
        ("inst_name",),
    )
    with pytest.raises(LaneBValidationError, match="列数不足"):
        parse_model_field_rows(rows, model_id="sample")


def test_LineProtocol到VM严格传播指标身份_值与纳秒时间():
    line = (
        r"sample_info,instance_id=cmdb_123456,model_id=sample,"
        r"inst_name=node.example.invalid value=7i 1700000000123000000"
    )
    vm_response = build_vm_response_from_line_protocol(line)
    row = vm_response["data"]["result"][0]
    assert row == {
        "metric": {
            "instance_id": "cmdb_123456",
            "model_id": "sample",
            "inst_name": "node.example.invalid",
            "__name__": "sample_info_value",
        },
        "value": [1700000000.123, "7"],
    }
    schema = build_case_vm_schema(vm_response, emitted_model_id="sample")
    target_metric_schema = schema["properties"]["data"]["properties"]["result"][
        "contains"
    ]["properties"]["metric"]["properties"]
    assert target_metric_schema["instance_id"] == {"const": "cmdb_123456"}
    validate(vm_response, schema)

    for field, drifted_value in (
        ("__name__", "wrong_info_value"),
        ("model_id", "wrong"),
        ("instance_id", "cmdb_999999"),
    ):
        drifted = json.loads(json.dumps(vm_response))
        drifted["data"]["result"][0]["metric"][field] = drifted_value
        with pytest.raises(ValidationError):
            validate(drifted, schema)

    invalid_time = json.loads(json.dumps(vm_response))
    invalid_time["data"]["result"][0]["value"][0] = 9_999_999_999
    with pytest.raises(ValidationError):
        validate(invalid_time, schema)


def test_LineProtocol到VM拒绝非纳秒时间与身份漂移():
    with pytest.raises(LaneBValidationError, match="19位纳秒"):
        build_vm_response_from_line_protocol(
            "sample,instance_id=cmdb-invalid value=1i 9999999999"
        )


def test_CMDB专属schema拒绝缺字段_错误重命名和类型漂移():
    expected = {
        "model_id": "sample",
        "instance_count_min": 1,
        "expected_instances": [
            {"inst_name": "sample-contract", "enabled": False, "capacity": 0}
        ],
        "optional_absent_fields": [],
        "review_basis": "独立静态审阅",
    }
    schema = build_case_cmdb_schema(expected)
    validate(expected, schema)

    for drifted in (
        {
            **expected,
            "expected_instances": [{"enabled": False, "capacity": 0}],
        },
        {
            **expected,
            "expected_instances": [
                {
                    "instance_name": "sample-contract",
                    "enabled": False,
                    "capacity": 0,
                }
            ],
        },
        {
            **expected,
            "expected_instances": [
                {
                    "inst_name": "sample-contract",
                    "enabled": "false",
                    "capacity": 0,
                }
            ],
        },
    ):
        with pytest.raises(ValidationError):
            validate(drifted, schema)


LANE_B_READY_ENTRIES = tuple(
    entry
    for entry in lane_b_entries()
    if not load_lane_b_evidence(entry.case_id).missing_files
)


def test_LaneB静态制品覆盖不得落后于LaneA_ready集合():
    lane_a_ready = {
        item.case_id
        for item in audit_lane_a_evidence().validation
        if item.status == "ready"
    }
    assert not (lane_a_ready & set(lane_b_incomplete())), lane_b_incomplete()


@pytest.mark.parametrize(
    "entry", LANE_B_READY_ENTRIES, ids=lambda entry: entry.case_id
)
def test_LaneB静态制品通过schema且模型身份一致(entry):
    evidence = load_lane_b_evidence(entry.case_id)
    evidence.validate()
    vm_response = evidence.read_json("04_vm_response.json")
    expected = evidence.read_json("05_expected_cmdb.json")

    assert vm_response["status"] == "success"
    assert vm_response["data"]["resultType"] == "vector"
    assert expected.get("source_contract_model_id", entry.emitted_model_id) == entry.emitted_model_id
    if expected.get("write_mode", "graph_entity") != "ipam_service":
        assert expected["expected_instances"]
        assert all(row["inst_name"] for row in expected["expected_instances"])
        assert "expected_instance_subset" not in expected
        assert isinstance(expected["optional_absent_fields"], list)


@pytest.mark.parametrize(
    "entry", LANE_B_READY_ENTRIES, ids=lambda entry: entry.case_id
)
def test_独立Golden关键字段符合生产模型反射与字段类型(entry):
    expected = load_lane_b_evidence(entry.case_id).read_json(
        "05_expected_cmdb.json"
    )
    if expected.get("write_mode") == "ipam_service":
        assert expected["expected_vm_metric_subset"]
        return

    definitions = _production_model_field_definitions(expected["model_id"])
    fields = {
        field: definition.attr_type
        for field, definition in definitions.items()
    }
    rows = expected["expected_instances"]
    unknown = {
        field
        for row in rows
        for field in row
        if field != "assos" and field not in fields
    }
    assert not unknown, f"{entry.case_id}: Golden 含生产模型未定义字段 {sorted(unknown)}"
    mismatches = {
        field: {"model_type": fields[field], "value": value}
        for row in rows
        for field, value in row.items()
        if field != "assos"
        and not _value_matches_model_type(value, fields[field])
    }
    assert not mismatches, f"{entry.case_id}: Golden 字段类型与生产模型不一致 {mismatches}"
    required_fields = {
        field
        for field, definition in definitions.items()
        if definition.is_required
    } - {"organization"}
    for row in rows:
        missing_required = required_fields - set(row)
        assert not missing_required, (
            f"{entry.case_id}: Golden 缺生产模型必填字段 "
            f"{sorted(missing_required)}"
        )
        empty_required = {
            field for field in required_fields if row[field] in (None, "")
        }
        assert not empty_required, (
            f"{entry.case_id}: Golden 必填字段为空 {sorted(empty_required)}"
        )
        invalid_enums = {
            field: row[field]
            for field, definition in definitions.items()
            if definition.enum_values
            and field in row
            and row[field] not in ("", None)
            and row[field] not in definition.enum_values
        }
        assert not invalid_enums, (
            f"{entry.case_id}: Golden enum 不在生产选项 {invalid_enums}"
        )


@pytest.mark.parametrize(
    "entry", LANE_B_READY_ENTRIES, ids=lambda entry: entry.case_id
)
def test_VM响应经过真实注册插件后匹配独立静态CMDB_Golden(entry, monkeypatch):
    evidence = load_lane_b_evidence(entry.case_id)
    vm_response = evidence.read_json("04_vm_response.json")
    expected = evidence.read_json("05_expected_cmdb.json")

    expected_mode = expected.get("write_mode", "graph_entity")
    if expected_mode == "ipam_service":
        applied_ip_rows = {}
        plugin, result, query = _run_real_plugin(
            entry, vm_response, monkeypatch, applied_ip_rows=applied_ip_rows
        )
        _assert_query_identity(plugin, query, vm_response)
        assert plugin.raw_data == vm_response["data"]["result"]
        assert result == {"ip": []}
        assert applied_ip_rows["task"] is plugin.get_collect_inst()
        assert any(
            all(row.get(field) == value for field, value in expected["expected_vm_metric_subset"].items())
            for row in applied_ip_rows["rows"]
        )
        return

    plugin, result, query = _run_real_plugin(entry, vm_response, monkeypatch)
    output_model_id = _output_model_id(entry, expected)
    rows = result[output_model_id]

    _assert_query_identity(plugin, query, vm_response)
    assert plugin.raw_data == vm_response["data"]["result"]
    assert len(rows) >= expected["instance_count_min"], result
    assert rows == expected["expected_instances"]
    actual_fields = set().union(*(row.keys() for row in rows))
    assert expected["optional_absent_fields"] == sorted(
        _target_mapping_fields(plugin, output_model_id) - actual_fields
    )
    target_vm_response = build_vm_response_from_line_protocol(
        (evidence.fixture_dir / "03_line_protocol.txt").read_text(
            encoding="utf-8"
        )
    )
    target_metrics = [
        row["metric"] for row in target_vm_response["data"]["result"]
    ]
    unexpectedly_dropped = {
        field
        for field in expected["optional_absent_fields"]
        if any(metric.get(field) not in (None, "") for metric in target_metrics)
    }
    assert not unexpectedly_dropped, (
        f"{entry.case_id}: VM 已携带但清洗丢失 mapped 字段 "
        f"{sorted(unexpectedly_dropped)}"
    )


def test_真实插件仍会过滤超过一天的VM数据(monkeypatch):
    entry = next(
        item for item in lane_b_entries() if item.case_id == "aliyun_bucket"
    )
    vm_response = load_lane_b_evidence(entry.case_id).read_json(
        "04_vm_response.json"
    )

    plugin, result, query = _run_real_plugin(
        entry,
        vm_response,
        monkeypatch,
        clock_offset_seconds=2 * 24 * 60 * 60,
    )

    _assert_query_identity(plugin, query, vm_response)
    assert result["aliyun_bucket"] == []


def test_expected必须是静态字面量且不得携带运行时生成声明():
    for entry in LANE_B_READY_ENTRIES:
        expected = load_lane_b_evidence(entry.case_id).read_json(
            "05_expected_cmdb.json"
        )
        assert "generated_at" not in expected
        assert "generated_from_mapping" not in expected
        if expected.get("write_mode") != "ipam_service":
            assert "expected_instances" in expected
            assert "expected_instance_subset" not in expected


def test_VM_Golden锁定成功状态_向量形态_标签_值与新鲜时间戳():
    for entry in LANE_B_READY_ENTRIES:
        vm_response = load_lane_b_evidence(entry.case_id).read_json(
            "04_vm_response.json"
        )
        rows = vm_response["data"]["result"]
        assert rows
        for row in rows:
            metric = row["metric"]
            timestamp, value = row["value"]
            assert metric["__name__"].endswith("_info_gauge")
            assert metric["collect_status"] == "success"
            assert metric["instance_id"]
            assert isinstance(timestamp, (int, float)) and timestamp > 0
            assert isinstance(value, str)


def test_LaneB生产对象不得用skip或xfail掩盖缺口():
    source = __file__
    text = open(source, encoding="utf-8").read()
    assert "pytest." + "skip(" not in text
    assert "pytest.mark." + "xfail" not in text


@pytest.mark.parametrize(
    "case_id", ["qcloud_filesystem", "qcloud_rocketmq"]
)
def test_云API缺失可选数值字段不会阻断其余字段清洗(case_id, monkeypatch):
    entry = next(item for item in lane_b_entries() if item.case_id == case_id)
    vm_response = load_lane_b_evidence(case_id).read_json("04_vm_response.json")

    _, result, _ = _run_real_plugin(entry, vm_response, monkeypatch)

    assert result[entry.emitted_model_id]
    assert result[entry.emitted_model_id][0]["inst_name"]
